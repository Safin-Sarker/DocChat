"""Excel extraction utilities."""

import logging
from typing import Any, Dict, List

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _rows_to_markdown(rows: List[List[str]]) -> str:
    """Convert rows to markdown text."""
    if not rows:
        return ""

    column_count = max(len(row) for row in rows)
    normalized_rows = [row + [""] * (column_count - len(row)) for row in rows]
    header = normalized_rows[0]
    body = normalized_rows[1:]

    header_line = "| " + " | ".join(cell or "" for cell in header) + " |"
    sep_line = "| " + " | ".join("---" for _ in header) + " |"
    body_lines = [
        "| " + " | ".join(cell or "" for cell in row) + " |"
        for row in body
    ]
    return "\n".join([header_line, sep_line] + body_lines)


class ExcelExtractor:
    """Extract sheet text and table-like content from XLSX files."""

    def extract_sheets(self, xlsx_path: str) -> List[Dict[str, Any]]:
        """Extract each worksheet into chunkable text blocks."""
        try:
            workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
        except Exception as exc:
            logger.error("XLSX load failed: %s", exc)
            return []

        sheets_out: List[Dict[str, Any]] = []
        for sheet_num, worksheet in enumerate(workbook.worksheets, start=1):
            rows: List[List[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                cells = ["" if value is None else str(value).strip() for value in row]
                if any(cells):
                    rows.append(cells)

            if not rows:
                continue

            markdown = _rows_to_markdown(rows)
            if not markdown:
                continue

            sheet_text = f"# Sheet: {worksheet.title}\n\n{markdown}"
            sheets_out.append({
                "page_num": sheet_num,
                "text": sheet_text,
                "sheet_name": worksheet.title,
            })

        return sheets_out

    def extract_tables(self, xlsx_path: str) -> List[Dict[str, Any]]:
        """Extract each worksheet as a table markdown entry."""
        try:
            workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
        except Exception as exc:
            logger.error("XLSX load failed for tables: %s", exc)
            return []

        tables_out: List[Dict[str, Any]] = []
        for sheet_index, worksheet in enumerate(workbook.worksheets):
            rows: List[List[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                cells = ["" if value is None else str(value).strip() for value in row]
                if any(cells):
                    rows.append(cells)

            markdown = _rows_to_markdown(rows)
            if not markdown:
                continue

            tables_out.append({
                "page": sheet_index + 1,
                "table_index": 0,
                "sheet_name": worksheet.title,
                "markdown": markdown,
            })

        return tables_out
