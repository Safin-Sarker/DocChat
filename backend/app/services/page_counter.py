"""Lightweight page counter — counts pages before expensive processing."""

import math
from pathlib import Path

from app.core.config import settings


def count_pages(file_path: str, file_type: str) -> int:
    """Return the estimated page count for a file.

    This runs *before* the full multimodal processor so it must be fast
    and must not import heavy ML libraries.
    """
    if file_type == "pdf":
        return _count_pdf(file_path)
    if file_type == "docx":
        return _count_docx(file_path)
    if file_type == "xlsx":
        return _count_xlsx(file_path)
    if file_type == "pptx":
        return _count_pptx(file_path)
    if file_type == "txt":
        return _count_txt(file_path)
    if file_type == "image":
        return 1
    return 1


def _count_pdf(file_path: str) -> int:
    from PyPDF2 import PdfReader

    return len(PdfReader(file_path).pages)


def _count_docx(file_path: str) -> int:
    from docx import Document

    doc = Document(file_path)
    paragraphs = len(doc.paragraphs)
    return max(1, math.ceil(paragraphs / 40))


def _count_xlsx(file_path: str) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    count = len(wb.sheetnames)
    wb.close()
    return max(1, count)


def _count_pptx(file_path: str) -> int:
    from pptx import Presentation

    return max(1, len(Presentation(file_path).slides))


def _count_txt(file_path: str) -> int:
    text = Path(file_path).read_text(encoding="utf-8", errors="ignore")
    return max(1, math.ceil(len(text) / settings.TEXT_MAX_SECTION_CHARS))
